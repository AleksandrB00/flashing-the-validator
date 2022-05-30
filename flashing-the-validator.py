import xml.etree.ElementTree as ET
import os
from openpyxl import load_workbook
def convert():
 os.chdir( r"C:\Users\a.babaev\Desktop\Программирование")
 wb=load_workbook(filename='Штрих-Карт_497_20220131.xlsx', read_only=True)
 ws = wb['Лист1']
 a=input("Номер валидатора:")
 for row_cells in ws:
  for cell in row_cells:
    if a == cell.value:
      string_row = (cell.row)
      driver = ws.cell(string_row,cell.column+4)
      tid = ws.cell(string_row,cell.column+6)
      code = ws.cell(string_row,cell.column+7)
      inn = ws.cell(string_row,cell.column+9)
      city= ws.cell(string_row,cell.column+16)
      surname = ws.cell(string_row,cell.column+10)
      name = ws.cell(string_row,cell.column+11)
      patronimyc = ws.cell(string_row,cell.column+12)
      telephone = ws.cell(string_row,cell.column+13)
      login = ws.cell(string_row,cell.column+14)
      password = ws.cell(string_row,cell.column+15)
      os.chdir( r"C:\Users\a.babaev\Desktop\Программирование\Sert_conf_keys_KO" )
      tree = ET.parse('Settings' + str(cell.value) + '.xml')
      root = tree.getroot()
      element = root[0]
      element.text = str(driver.value)
      element2 = root[1]
      element2.text ="г." + " " + str(city.value) + " " + "ИНН" + " " + str(inn.value)
      element3 = root[7]
      element3.text= str(tid.value)
      element4 = root[8]
      element4.text = str(code.value)
      element5 = root[9]
      norm=int(element5.text)+1
      element5.text = str(norm)
      element11 = root[21]
      element11.text = []
      element6 = root[22]
      element6.text = str(inn.value)
      element7 = root[23]
      element7.text = str(name.value)
      element8 = root[24]
      element8.text = str(surname.value)
      element9 = root[25]
      element9.text = str(patronimyc.value)
      element10 = root[26]
      element10.text = str(telephone.value)
      element12 = root[13]
      element12.text = str(login.value)
      element13 = root[14]
      element13.text = str(password.value)
      tree.write('Settings' + str(cell.value) + '.xml', encoding='UTF-8')
      os.chdir( r"save_path")
      tree.write(str(cell.value) + '.xml', encoding='UTF-8')
convert()
while True:
  convert()
